from tkinter import *

import pandas as pd
import openpyxl,xlrd
from openpyxl import Workbook
import pathlib


file = pathlib.Path("Backened_Data.xlsx")
if file.exists ():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet["A1"]="Full Name"
    sheet["B1"]="Email"
    sheet["C1"]="Gender"
    sheet["D1"]="Country"
    sheet["E1"]="Programming"

    file.save("Backened_Data.xlsx")




def submit_field():
    y=entry_1.get()
    z=entry_2.get()
    print(y)
    print(z)
    
    if var.get()==1:
        gen="Male"
        print("Male")
    else:
        print("female")
        
    country=c.get()
    print(country)

    
    file=openpyxl.load_workbook("Backened_Data.xlsx")
    sheet=file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=y)
    sheet.cell(column=2,row=sheet.max_row,value=z)
    if var.get()==1:
        gen="Male"
        print("Male")
        sheet.cell(column=3,row=sheet.max_row,value="Male")
    else:
        print("female")
        sheet.cell(column=3,row=sheet.max_row,value="Female")
    sheet.cell(column=4,row=sheet.max_row,value=country)
    if var1.get() ==1:
        print("Java")
        sheet.cell(column=5,row=sheet.max_row,value="Java") 
    if var2.get() ==1:
        print("Python")
        sheet.cell(column=5,row=sheet.max_row,value="Python")         
    file.save("Backened_Data.xlsx")

    xlfile = pd.read_excel('Backened_Data.xlsx', 'Sheet') # reading xl file 
    xlfile.to_csv('Backened_Data.csv', index=False)#conversion to csv

    
root = Tk()
root.geometry('500x500')
root.title("Registration Form")
label_0 = Label(root, text="Registration form",width=20,font=("bold", 20))
label_0.place(x=90,y=53)


label_1 = Label(root, text="FullName",width=20,font=("bold", 10))
label_1.place(x=80,y=130)

entry_1 = Entry(root)
entry_1.place(x=240,y=130)

label_2 = Label(root, text="Email",width=20,font=("bold", 10))
label_2.place(x=68,y=180)

entry_2 = Entry(root)
entry_2.place(x=240,y=180)

label_3 = Label(root, text="Gender",width=20,font=("bold", 10))
label_3.place(x=70,y=230)
var = IntVar()
Radiobutton(root, text="Male",padx = 5, variable=var, value=1).place(x=235,y=230)
Radiobutton(root, text="Female",padx = 20, variable=var, value=2).place(x=290,y=230)

label_4 = Label(root, text="country",width=20,font=("bold", 10))
label_4.place(x=70,y=280)

list1 = ['Canada','India','UK','USA','Australia','South Africa'];
c=StringVar()
droplist=OptionMenu(root,c, *list1)
droplist.config(width=15)
c.set('select your country') 
droplist.place(x=240,y=280)

label_4 = Label(root, text="Programming",width=20,font=("bold", 10))
label_4.place(x=85,y=330)
var1 = IntVar()
Checkbutton(root, text="java", variable=var1).place(x=235,y=330)
var2 = IntVar()
Checkbutton(root, text="python", variable=var2).place(x=290,y=330)

#Button(root, text='Submit',width=20,bg='blue',fg='white').place(x=180,y=380)
Button(root, text='Submit',command=submit_field,width=20,bg='blue',fg='white').place(x=180,y=380)

root.mainloop()

